from playwright.sync_api import sync_playwright
import time
import random
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os

def random_sleep(a=1, b=3):
    time.sleep(random.uniform(a, b))

def slow_type(page, selector, text, delay=0.1):
    page.click(selector)
    for char in text:
        page.keyboard.type(char)
        time.sleep(delay)

def style_excel(input_file, output_file):
    # Make sure the file exists
    if not os.path.exists(input_file):
        print("Results.xlsx not found.")
        return
    wb = load_workbook(input_file)
    ws = wb.active
    header = ws[1]
    for cell in header:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 4
        ws.column_dimensions[column].width = adjusted_width
    wb.save(output_file)
    print(f"{output_file} created successfully!")

job_titles = []

with sync_playwright() as p:
    browser = p.chromium.launch(headless=False, slow_mo=50)
    page = browser.new_page()
    page.goto('https://www.adzuna.com/')
    random_sleep(2, 4)
    page.wait_for_selector('header nav ul li:first-child a', state="visible")
    page.click('header nav ul li:first-child a')
    page.wait_for_load_state('domcontentloaded')
    random_sleep(1, 3)
    page.wait_for_selector('#email', state='visible')
    random_sleep(1, 4)
    slow_type(page, '#email', 'kalabiq1@gmail.com', delay=0.15)
    random_sleep(2, 4)
    slow_type(page, '#password', 'A55975597a!', delay=0.2)
    random_sleep(1, 5)
    page.mouse.wheel(0, 300)
    random_sleep(2, 4)
    button = page.query_selector('form button[type="submit"]')
    box = button.bounding_box()
    page.mouse.move(box['x'] + box['width']/2, box['y'] + box['height']/2, steps=25)
    random_sleep(1, 2)
    page.get_by_role('button', name='Login').click()
    page.wait_for_load_state('domcontentloaded')
    random_sleep(3, 6)
    page.wait_for_selector('#what', state='visible')
    slow_type(page, '#what', 'Logistics', delay=0.3)
    random_sleep(1, 2)
    page.click('button[type="submit"]')
    random_sleep(3, 5)

    while True:
        page.wait_for_selector('div.flex.gap-4 h2 a')
        links = page.query_selector_all('div.flex.gap-4 h2 a')
        for link in links:
            text = link.inner_text().strip()
            job_titles.append(text)
            print(text)
        try:
            next_button = page.query_selector('a.flex-auto.lg\\:inline-block.px-3.leading-10.border.border-solid.border-adzuna-green-500.text-adzuna-green-500.rounded-lg.hover\\:text-white.hover\\:bg-adzuna-green-500.md\\:ml-1')
            if next_button:
                next_button.click()
                random_sleep(2, 4)
            else:
                print("No more pages")
                break
        except Exception as e:
            print("Error")
            break

    for job_title in job_titles:
        print(f"Testing job title: {job_title}")
        try:
            # Wait for the job title link to be visible
            job_locator = page.locator(f"text={job_title}")
            job_locator.wait_for(state="visible", timeout=5000)  # Adding timeout to avoid hanging
            print(f"Clicking job title: {job_title}")
            job_locator.click()  # Click on the job title link
            random_sleep(3, 5)  # Stay on the job page for a few seconds
            page.go_back()  # Go back to the search results page
            random_sleep(3, 4)  # Wait before clicking the next job title
        except Exception as e:
            print(f"Failed to test {job_title}: {str(e)}")

    # Save the data in a styled Excel file
    df = pd.DataFrame(job_titles, columns=["Job title"])
    df.to_excel('Styled_Results.xlsx', index=False)
    style_excel('Styled_Results.xlsx', 'Styled_Results.xlsx')
    print("Scraped data saved to Styled_Results.xlsx!")

    browser.close()
