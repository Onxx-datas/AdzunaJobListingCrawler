from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os





input_file = "Results.xlsx"
output_file = "Styled_Results.xlsx"





if not os.path.exists(input_file):
    print("Results.xlsx not found.")
    exit()





wb = load_workbook(input_file)
ws = wb.active
header = ws[1]





for cell in header:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(horizontal="left", vertical="center")
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = max_length + 4
    ws.column_dimensions[column].width = adjusted_width





wb.save(output_file)
print("Styled_Results.xlsx created successfully!")
